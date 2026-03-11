import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.modules.profitability.model import ProfitabilityData, ProfitabilityCalculator
from app.file_utils import get_download_path, get_timestamped_filename

# 색상
BLUE_FILL   = PatternFill("solid", fgColor="0D3B5E")
HEADER_FILL = PatternFill("solid", fgColor="EEF2FF")
LIGHT_FILL  = PatternFill("solid", fgColor="F8FAFF")
GREEN_FILL  = PatternFill("solid", fgColor="D1FAE5")
RED_FILL    = PatternFill("solid", fgColor="FEE2E2")

WHITE_FONT  = Font(color="FFFFFF", bold=True, name="맑은 고딕")
BOLD_FONT   = Font(bold=True, name="맑은 고딕")
NORMAL_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT  = Font(bold=True, size=14, name="맑은 고딕", color="0D3B5E")

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _cell(ws, row, col, value, font=None, fill=None, number_format=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    if alignment:
        c.alignment = alignment
    c.border = THIN_BORDER
    return c


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def export_xlsx(data: ProfitabilityData) -> str:
    calc = ProfitabilityCalculator(data)
    wb = openpyxl.Workbook()

    _build_summary_sheet(wb.active, data, calc)
    wb.active.title = "요약 대시보드"

    cost_sheet = wb.create_sheet("비용 분석")
    _build_cost_sheet(cost_sheet, data, calc)

    rev_sheet = wb.create_sheet("수익 분석")
    _build_revenue_sheet(rev_sheet, data, calc)

    scenario_sheet = wb.create_sheet("시나리오 분석")
    _build_scenario_sheet(scenario_sheet, data, calc)

    file_name = get_timestamped_filename("수익성분석", "xlsx")
    path = get_download_path(file_name)
    wb.save(path)
    return path


def _build_summary_sheet(ws, data: ProfitabilityData, calc: ProfitabilityCalculator):
    _set_col_widths(ws, {"A": 22, "B": 18, "C": 18, "D": 18})

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"수익성 분석 요약 — {data.project_name}"
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["지표", "값", "비고", ""]
    for col, h in enumerate(headers, 1):
        _cell(ws, 2, col, h, font=WHITE_FONT, fill=BLUE_FILL,
              alignment=Alignment(horizontal="center"))

    kv_rows = [
        ("프로젝트명",      data.project_name,                           ""),
        ("분석 기간",       f"{data.period_year}년",                     ""),
        ("할인율",          f"{data.discount_rate}%",                     ""),
        ("",               "",                                            ""),
        ("총 수익",         calc.format_currency(calc.total_revenue()),    ""),
        ("고정비",          calc.format_currency(calc.total_fixed_cost()), ""),
        ("변동비",          calc.format_currency(calc.total_variable_cost()), ""),
        ("총 비용",         calc.format_currency(calc.total_cost()),       ""),
        ("",               "",                                            ""),
        ("순이익",          calc.format_currency(calc.net_profit()),       "흑자" if calc.net_profit() >= 0 else "적자"),
        ("ROI",          f"{calc.profit_margin():.1f}%",               ""),
        ("손익분기점",      _fmt_bep(calc.breakeven_years()),             ""),
    ]

    for i, (k, v, note) in enumerate(kv_rows, 3):
        fill = LIGHT_FILL if i % 2 == 0 else None
        if k:
            _cell(ws, i, 1, k, font=BOLD_FONT, fill=HEADER_FILL)
            profit_fill = GREEN_FILL if k == "순이익" and calc.net_profit() >= 0 else (RED_FILL if k == "순이익" else fill)
            _cell(ws, i, 2, v, fill=profit_fill, font=BOLD_FONT if k == "순이익" else NORMAL_FONT)
            _cell(ws, i, 3, note, font=NORMAL_FONT, fill=fill)
            _cell(ws, i, 4, "", fill=fill)
        else:
            for col in range(1, 5):
                _cell(ws, i, col, "")


def _build_cost_sheet(ws, data: ProfitabilityData, calc: ProfitabilityCalculator):
    _set_col_widths(ws, {"A": 24, "B": 12, "C": 16, "D": 14, "E": 24})
    ws["A1"].value = "비용 분석"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    headers = ["비용 항목", "유형", "금액", "주기", "비고"]
    for col, h in enumerate(headers, 1):
        _cell(ws, 2, col, h, font=WHITE_FONT, fill=BLUE_FILL,
              alignment=Alignment(horizontal="center"))

    curr_fmt = "#,##0" if data.currency == "KRW" else "#,##0.00"
    for i, item in enumerate(data.cost_items, 3):
        fill = LIGHT_FILL if i % 2 == 0 else None
        _cell(ws, i, 1, item.name,     font=NORMAL_FONT, fill=fill)
        _cell(ws, i, 2, item.category, font=NORMAL_FONT, fill=fill)
        _cell(ws, i, 3, item.amount,   font=NORMAL_FONT, fill=fill, number_format=curr_fmt)
        _cell(ws, i, 4, item.period,   font=NORMAL_FONT, fill=fill)
        _cell(ws, i, 5, item.note,     font=NORMAL_FONT, fill=fill)

    row = len(data.cost_items) + 3
    _cell(ws, row, 1, "합계 (기간 전체)", font=BOLD_FONT, fill=HEADER_FILL)
    _cell(ws, row, 3, calc.total_cost(), font=BOLD_FONT, fill=HEADER_FILL, number_format=curr_fmt)


def _build_revenue_sheet(ws, data: ProfitabilityData, calc: ProfitabilityCalculator):
    _set_col_widths(ws, {"A": 24, "B": 16, "C": 14, "D": 18, "E": 24})
    ws["A1"].value = "수익 분석"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    headers = ["수익원", "단가", "수량", "합계", "비고"]
    for col, h in enumerate(headers, 1):
        _cell(ws, 2, col, h, font=WHITE_FONT, fill=BLUE_FILL,
              alignment=Alignment(horizontal="center"))

    curr_fmt = "#,##0" if data.currency == "KRW" else "#,##0.00"
    for i, item in enumerate(data.revenue_items, 3):
        fill = LIGHT_FILL if i % 2 == 0 else None
        _cell(ws, i, 1, item.name,        font=NORMAL_FONT, fill=fill)
        _cell(ws, i, 2, item.unit_price,  font=NORMAL_FONT, fill=fill, number_format=curr_fmt)
        _cell(ws, i, 3, item.quantity,    font=NORMAL_FONT, fill=fill)
        _cell(ws, i, 4, item.total,       font=NORMAL_FONT, fill=fill, number_format=curr_fmt)
        _cell(ws, i, 5, item.note,        font=NORMAL_FONT, fill=fill)

    row = len(data.revenue_items) + 3
    _cell(ws, row, 1, "총 수익", font=BOLD_FONT, fill=GREEN_FILL)
    _cell(ws, row, 4, calc.total_revenue(), font=BOLD_FONT, fill=GREEN_FILL, number_format=curr_fmt)


def _build_scenario_sheet(ws, data: ProfitabilityData, calc: ProfitabilityCalculator):
    _set_col_widths(ws, {"A": 22, "B": 18, "C": 18, "D": 18})
    ws["A1"].value = "시나리오 분석"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    headers = ["지표", "낙관적", "기본", "비관적"]
    for col, h in enumerate(headers, 1):
        _cell(ws, 2, col, h, font=WHITE_FONT, fill=BLUE_FILL,
              alignment=Alignment(horizontal="center"))

    optimistic = calc.scenario(data.optimistic_revenue_delta, data.optimistic_cost_delta)
    base       = calc.scenario(0, 0)
    pessimistic = calc.scenario(data.pessimistic_revenue_delta, data.pessimistic_cost_delta)

    curr_fmt = "#,##0" if data.currency == "KRW" else "#,##0.00"
    rows = [
        ("총 수익",  optimistic["revenue"],  base["revenue"],  pessimistic["revenue"],  curr_fmt),
        ("총 비용",  optimistic["cost"],     base["cost"],     pessimistic["cost"],     curr_fmt),
        ("순이익",   optimistic["profit"],   base["profit"],   pessimistic["profit"],   curr_fmt),
        ("ROI (%)",  optimistic["roi"],      base["roi"],      pessimistic["roi"],      "0.0%"),
    ]

    for i, (label, opt, base_val, pess, fmt) in enumerate(rows, 3):
        fill = LIGHT_FILL if i % 2 == 0 else None
        _cell(ws, i, 1, label,    font=BOLD_FONT,  fill=HEADER_FILL)
        _cell(ws, i, 2, opt,      font=NORMAL_FONT, fill=GREEN_FILL, number_format=fmt)
        _cell(ws, i, 3, base_val, font=NORMAL_FONT, fill=fill,       number_format=fmt)
        _cell(ws, i, 4, pess,     font=NORMAL_FONT, fill=RED_FILL,   number_format=fmt)

    row = len(rows) + 4
    scenario_labels = [
        ("수익 변동률", f"+{data.optimistic_revenue_delta}%", "0%", f"{data.pessimistic_revenue_delta}%"),
        ("비용 변동률", f"{data.optimistic_cost_delta}%",     "0%", f"+{data.pessimistic_cost_delta}%"),
    ]
    for j, (label, opt, base_str, pess) in enumerate(scenario_labels):
        _cell(ws, row + j, 1, label,    font=NORMAL_FONT, fill=HEADER_FILL)
        _cell(ws, row + j, 2, opt,      font=NORMAL_FONT)
        _cell(ws, row + j, 3, base_str, font=NORMAL_FONT)
        _cell(ws, row + j, 4, pess,     font=NORMAL_FONT)


def _fmt_bep(years: float) -> str:
    if years == float("inf"):
        return "손익분기 도달 불가"
    return f"{years:.1f}년"
